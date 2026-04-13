from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
import os, io, base64, struct, zlib
from datetime import datetime
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'activos-colegio-2025-secret')
DATABASE_URL = os.environ.get('DATABASE_URL', '')
FICHA_PIN = os.environ.get('FICHA_PIN', '1234')

TIPOS = [
    'Equipamiento Tecnológico',
    'Equipamiento Audiovisual',
    'Equipamiento Deportivo',
    'Mobiliario',
    'Muebles y Útiles',
    'Otro',
]

# SII vida util por tipo
VIDA_UTIL_SII = {
    'Equipamiento Tecnológico': 6,
    'Equipamiento Audiovisual': 7,
    'Equipamiento Deportivo':   5,
    'Mobiliario':               7,
    'Muebles y Útiles':         7,
    'Otro':                     7,
}

def get_db():
    if DATABASE_URL:
        import psycopg2, psycopg2.extras
        return psycopg2.connect(DATABASE_URL), 'pg'
    import sqlite3
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'activos.db')
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn, 'sqlite'

def db_fetchall(sql, params=()):
    conn, mode = get_db()
    if mode == 'pg':
        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql.replace('?','%s'), params)
    else:
        cur = conn.cursor()
        cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def db_fetchone(sql, params=()):
    conn, mode = get_db()
    if mode == 'pg':
        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql.replace('?','%s'), params)
    else:
        cur = conn.cursor()
        cur.execute(sql, params)
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None

def db_execute(sql, params=(), commit=True):
    conn, mode = get_db()
    if mode == 'pg':
        cur = conn.cursor()
        cur.execute(sql.replace('?','%s'), params)
    else:
        cur = conn.cursor()
        cur.execute(sql, params)
    if commit:
        conn.commit()
    conn.close()

def init_db():
    conn, mode = get_db()
    if mode == 'pg':
        import psycopg2.extras
        cur = conn.cursor()
        cur.execute('''CREATE TABLE IF NOT EXISTS activos (
            id TEXT PRIMARY KEY,
            tipo TEXT, subtipo TEXT, marca TEXT, modelo TEXT, serie TEXT,
            estado TEXT, edificio TEXT, sala TEXT, responsable TEXT,
            fecha_compra TEXT, precio REAL, documento TEXT, vida_util INTEGER,
            observaciones TEXT, foto TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        cur.execute('''CREATE TABLE IF NOT EXISTS movimientos (
            id SERIAL PRIMARY KEY,
            activo_id TEXT, tipo TEXT, descripcion TEXT,
            usuario TEXT, fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        cur.execute('''CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            nombre TEXT, email TEXT UNIQUE, password TEXT, rol TEXT DEFAULT \'consulta\'
        )''')
        admin_email = os.environ.get('ADMIN_EMAIL','admin@colegio.cl')
        admin_pass  = os.environ.get('ADMIN_PASS','admin123')
        cur.execute("INSERT INTO usuarios (nombre,email,password,rol) VALUES (%s,%s,%s,%s) ON CONFLICT (email) DO NOTHING",
                    ('Administrador', admin_email, admin_pass, 'admin'))
        cur.execute("UPDATE usuarios SET password=%s, email=%s WHERE rol='admin'",
                    (admin_pass, admin_email))
    else:
        cur = conn.cursor()
        cur.execute('''CREATE TABLE IF NOT EXISTS activos (
            id TEXT PRIMARY KEY,
            tipo TEXT, subtipo TEXT, marca TEXT, modelo TEXT, serie TEXT,
            estado TEXT, edificio TEXT, sala TEXT, responsable TEXT,
            fecha_compra TEXT, precio REAL, documento TEXT, vida_util INTEGER,
            observaciones TEXT, foto TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )''')
        try: cur.execute("ALTER TABLE activos ADD COLUMN foto TEXT")
        except: pass
        cur.execute('''CREATE TABLE IF NOT EXISTS movimientos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            activo_id TEXT, tipo TEXT, descripcion TEXT,
            usuario TEXT, fecha TEXT DEFAULT CURRENT_TIMESTAMP
        )''')
        cur.execute('''CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT, email TEXT UNIQUE, password TEXT, rol TEXT DEFAULT \'consulta\'
        )''')
        admin_email = os.environ.get('ADMIN_EMAIL','admin@colegio.cl')
        admin_pass  = os.environ.get('ADMIN_PASS','admin123')
        cur.execute("INSERT OR IGNORE INTO usuarios (nombre,email,password,rol) VALUES (?,?,?,?)",
                    ('Administrador', admin_email, admin_pass, 'admin'))
        cur.execute("UPDATE usuarios SET password=?, email=? WHERE rol='admin'",
                    (admin_pass, admin_email))
    conn.commit()
    conn.close()

CENTROS_COSTO = ['Dirección', 'Unidad Técnica Pedagógica (UTP)', 'Inspectoría General', 'Administración y Finanzas', 'Educación Básica', 'Educación Media', 'Educación Parvularia', 'Orientación', 'Convivencia Escolar', 'Biblioteca / CRA', 'Laboratorio de Ciencias', 'Sala de Computación', 'Educación Física / Deportes', 'PIE (Programa Integración Escolar)']

with app.app_context():
    init_db()
    # Migracion: agregar centro_costo (commit propio para que no falle silenciosamente)
    try:
        conn_cc, mode_cc = get_db()
        cur_cc = conn_cc.cursor()
        if mode_cc == 'pg':
            cur_cc.execute("ALTER TABLE activos ADD COLUMN IF NOT EXISTS centro_costo TEXT DEFAULT ''")
        else:
            try:
                cur_cc.execute("ALTER TABLE activos ADD COLUMN centro_costo TEXT DEFAULT ''")
            except: pass
        conn_cc.commit()
        conn_cc.close()
    except Exception as e_cc:
        print(f"Migracion centro_costo: {e_cc}")

    # Migracion: crear tabla mantenciones si no existe
    try:
        conn_m, mode_m = get_db()
        cur_m = conn_m.cursor()
        # centro_costo ya migrado arriba
        if mode_m == 'pg':
            cur_m.execute('''CREATE TABLE IF NOT EXISTS movimientos_activos (
                id SERIAL PRIMARY KEY,
                activo_id TEXT NOT NULL,
                tipo TEXT DEFAULT 'entrada',
                descripcion TEXT,
                edificio TEXT,
                responsable TEXT,
                usuario TEXT,
                fecha TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')
            cur_m.execute('''CREATE TABLE IF NOT EXISTS mantenciones (
                id SERIAL PRIMARY KEY,
                activo_id TEXT NOT NULL,
                fecha TEXT,
                tipo TEXT DEFAULT 'correctiva',
                descripcion TEXT,
                costo REAL DEFAULT 0,
                proveedor TEXT,
                estado TEXT DEFAULT 'solucionado',
                proxima_fecha TEXT,
                usuario TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')
        else:
            cur_m.execute('''CREATE TABLE IF NOT EXISTS movimientos_activos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                activo_id TEXT NOT NULL,
                tipo TEXT DEFAULT 'entrada',
                descripcion TEXT,
                edificio TEXT,
                responsable TEXT,
                usuario TEXT,
                fecha TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )''')
            cur_m.execute('''CREATE TABLE IF NOT EXISTS mantenciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                activo_id TEXT NOT NULL,
                fecha TEXT,
                tipo TEXT DEFAULT 'correctiva',
                descripcion TEXT,
                costo REAL DEFAULT 0,
                proveedor TEXT,
                estado TEXT DEFAULT 'solucionado',
                proxima_fecha TEXT,
                usuario TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )''')
        conn_m.commit()
        conn_m.close()
    except Exception as e_m:
        print(f"Migracion mantenciones: {e_m}")

def next_id(fecha_compra=None):
    # Determinar año de compra para el prefijo
    year = datetime.now().year % 100
    if fecha_compra:
        try:
            for fmt in ['%d-%m-%Y','%Y-%m-%d','%d/%m/%Y']:
                try:
                    year = datetime.strptime(str(fecha_compra)[:10], fmt).year % 100
                    break
                except: pass
        except: pass
    # Correlativo GLOBAL — buscar el mayor en TODOS los activos AF-XX-XXXX
    rows = db_fetchall("SELECT id FROM activos WHERE id LIKE 'AF-%'")
    nums = []
    for r in rows:
        try:
            parts = str(r['id']).split('-')
            if len(parts) >= 3:
                nums.append(int(parts[-1]))
        except: pass
    nxt = max(nums) + 1 if nums else 1000
    return f"AF-{year:02d}-{nxt}"

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session: return redirect('/login')
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session: return redirect('/login')
        if session.get('rol') != 'admin': return jsonify({'error':'Sin permisos'}), 403
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET','POST'])
def login():
    error = ''
    if request.method == 'POST':
        email = request.form.get('email','').strip()
        password = request.form.get('password','')
        u = db_fetchone("SELECT * FROM usuarios WHERE email=? AND password=?", (email, password))
        if u:
            session['user'] = u['nombre']
            session['rol'] = u['rol']
            session['email'] = u['email']
            return redirect('/')
        error = 'Email o contraseña incorrectos'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

@app.route('/')
@login_required
def index():
    return render_template('index.html', user=session['user'], rol=session['rol'], tipos=TIPOS, centros_costo=CENTROS_COSTO, edificios=['Básica','Media','Parvularia','Administración'])

@app.route('/api/activos', methods=['GET'])
@login_required
def get_activos():
    q = request.args.get('q','')
    tipo = request.args.get('tipo','')
    estado = request.args.get('estado','')
    edificio = request.args.get('edificio','')
    sql = "SELECT * FROM activos WHERE 1=1"
    params = []
    if q:
        sql += " AND (id LIKE ? OR marca LIKE ? OR modelo LIKE ? OR serie LIKE ? OR responsable LIKE ?)"
        p = f'%{q}%'; params += [p,p,p,p,p]
    if tipo:     sql += " AND tipo=?";     params.append(tipo)
    if estado:   sql += " AND estado=?";   params.append(estado)
    if edificio: sql += " AND edificio=?"; params.append(edificio)
    sql += " ORDER BY id DESC"
    return jsonify(db_fetchall(sql, params))

@app.route('/api/activos/<id>', methods=['GET'])
@login_required
def get_activo(id):
    a = db_fetchone("SELECT * FROM activos WHERE id=?", (id,))
    if not a: return jsonify({'error':'No encontrado'}), 404
    movs = db_fetchall("SELECT * FROM movimientos WHERE activo_id=? ORDER BY fecha DESC", (id,))
    # Calcular depreciacion
    dep = calcular_depreciacion(a)
    return jsonify({'activo': a, 'movimientos': movs, 'depreciacion': dep})

def calcular_depreciacion(a):
    try:
        if not a.get('fecha_compra') or not a.get('precio'):
            return None
        fecha_str = str(a['fecha_compra'])
        # Intentar parsear fecha en varios formatos
        for fmt in ['%d-%m-%Y','%Y-%m-%d','%d/%m/%Y']:
            try:
                fecha = datetime.strptime(fecha_str[:10], fmt)
                break
            except: fecha = None
        if not fecha: return None

        precio_original = float(a['precio'])
        if precio_original <= 0: return None

        tipo = a.get('tipo','Otro')
        vida_util = int(a.get('vida_util') or VIDA_UTIL_SII.get(tipo, 7))
        tasa_anual = 1.0 / vida_util

        hoy = datetime.now()
        anos_transcurridos = (hoy - fecha).days / 365.25
        valor_residual = precio_original * 0.10  # 10% valor residual SII

        depreciacion_acumulada = min(precio_original - valor_residual,
                                     (precio_original - valor_residual) * tasa_anual * anos_transcurridos)
        valor_actual = max(valor_residual, precio_original - depreciacion_acumulada)
        porcentaje_dep = min(100, (depreciacion_acumulada / (precio_original - valor_residual)) * 100) if precio_original > valor_residual else 100

        fecha_termino = datetime(fecha.year + vida_util, fecha.month, fecha.day)
        anos_restantes = max(0, (fecha_termino - hoy).days / 365.25)

        return {
            'precio_original':      round(precio_original),
            'valor_actual':         round(valor_actual),
            'valor_residual':       round(valor_residual),
            'depreciacion_acum':    round(depreciacion_acumulada),
            'porcentaje_dep':       round(porcentaje_dep, 1),
            'anos_transcurridos':   round(anos_transcurridos, 1),
            'anos_restantes':       round(anos_restantes, 1),
            'vida_util':            vida_util,
            'tasa_anual':           round(tasa_anual * 100, 2),
            'fecha_termino':        fecha_termino.strftime('%d-%m-%Y'),
            'estado_dep':           'Depreciado' if porcentaje_dep >= 100 else
                                    'Crítico' if porcentaje_dep >= 80 else
                                    'Avanzado' if porcentaje_dep >= 50 else 'Normal',
        }
    except Exception as e:
        return None

@app.route('/api/activos', methods=['POST'])
@admin_required
def crear_activo():
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No se recibieron datos'}), 400
        aid = next_id(data.get('fecha_compra',''))
        vida = data.get('vida_util') or VIDA_UTIL_SII.get(data.get('tipo','Otro'), 7)
        centro_costo = data.get('centro_costo') or ''
        # INSERT sin centro_costo para compatibilidad con BD que no tiene la columna aún
        db_execute('''INSERT INTO activos
            (id,tipo,subtipo,marca,modelo,serie,estado,edificio,sala,responsable,
             fecha_compra,precio,documento,vida_util,observaciones,foto)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (aid, data.get('tipo'), data.get('subtipo'), data.get('marca'),
             data.get('modelo'), data.get('serie'), data.get('estado','Bueno'),
             data.get('edificio'), data.get('sala'), data.get('responsable'),
             data.get('fecha_compra'), data.get('precio',0), data.get('documento'),
             vida, data.get('observaciones',''), data.get('foto','')))
        # UPDATE separado para centro_costo (columna puede haberse agregado via migración)
        try:
            db_execute("UPDATE activos SET centro_costo=? WHERE id=?", (centro_costo, aid))
        except Exception as e_cc:
            print(f"centro_costo update omitido: {e_cc}")
        db_execute("INSERT INTO movimientos (activo_id,tipo,descripcion,usuario) VALUES (?,?,?,?)",
                   (aid,'Alta','Activo registrado en el sistema',session['user']))
        return jsonify({'id': aid, 'ok': True})
    except Exception as e:
        print(f"Error crear_activo: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/activos/<id>', methods=['PUT'])
@admin_required
def editar_activo(id):
    data = request.json
    old = db_fetchone("SELECT * FROM activos WHERE id=?", (id,))
    if not old: return jsonify({'error':'No encontrado'}), 404
    campos = ['tipo','subtipo','marca','modelo','serie','estado','edificio','sala',
              'responsable','fecha_compra','precio','documento','vida_util','observaciones','foto','centro_costo']
    updates = {c: data[c] for c in campos if c in data}
    if updates:
        sets = ', '.join(f"{c}=?" for c in updates)
        db_execute(f"UPDATE activos SET {sets} WHERE id=?", list(updates.values())+[id])
    cambios = []
    for c in ['estado','edificio','sala','responsable']:
        if c in data and str(old.get(c,'')) != str(data[c]):
            cambios.append(f"{c}: '{old.get(c)}' → '{data[c]}'")
    if cambios:
        db_execute("INSERT INTO movimientos (activo_id,tipo,descripcion,usuario) VALUES (?,?,?,?)",
                   (id,'Edición',' | '.join(cambios),session['user']))
    return jsonify({'ok': True})

@app.route('/api/activos/<id>', methods=['DELETE'])
@admin_required
def eliminar_activo(id):
    if not db_fetchone("SELECT id FROM activos WHERE id=?", (id,)):
        return jsonify({'error':'No encontrado'}), 404
    db_execute("DELETE FROM movimientos WHERE activo_id=?", (id,))
    db_execute("DELETE FROM activos WHERE id=?", (id,))
    return jsonify({'ok': True})

@app.route('/api/activos/<id>/traslado', methods=['POST'])
@login_required
def traslado(id):
    data = request.json
    old = db_fetchone("SELECT edificio,sala,responsable,centro_costo FROM activos WHERE id=?", (id,))
    new_edificio   = data.get('edificio', old['edificio'] or '')
    new_sala       = data.get('sala', old['sala'] or '')
    new_resp       = data.get('responsable', old['responsable'] or '')
    new_cc         = data.get('centro_costo', old.get('centro_costo','') or '')

    # Descripcion detallada del traslado
    from_ubicacion = f"{old['edificio'] or '—'} — {old['sala'] or '—'}"
    to_ubicacion   = f"{new_edificio} — {new_sala}"
    old_resp       = old['responsable'] or '—'
    old_cc         = old.get('centro_costo','') or '—'
    new_cc_label   = new_cc or '—'

    desc = (f"DESDE: {from_ubicacion} | HACIA: {to_ubicacion} | "
            f"Responsable anterior: {old_resp} → Nuevo: {new_resp} | "
            f"Centro costo: {old_cc} → {new_cc_label}")

    db_execute("UPDATE activos SET edificio=?,sala=?,responsable=?,centro_costo=? WHERE id=?",
               (new_edificio, new_sala, new_resp, new_cc, id))
    db_execute("INSERT INTO movimientos (activo_id,tipo,descripcion,usuario) VALUES (?,?,?,?)",
               (id, 'Traslado', desc, session['user']))
    return jsonify({'ok': True})

@app.route('/api/activos/<id>/foto', methods=['POST'])
@admin_required
def subir_foto(id):
    if 'foto' not in request.files: return jsonify({'error':'No se envió archivo'}), 400
    file = request.files['foto']
    ext = file.filename.rsplit('.',1)[-1].lower()
    if ext not in {'jpg','jpeg','png','gif','webp'}: return jsonify({'error':'Formato no permitido'}), 400
    data = file.read()
    if len(data) > 5*1024*1024: return jsonify({'error':'Imagen muy grande (máx 5MB)'}), 400
    b64 = f"data:image/{ext};base64,"+base64.b64encode(data).decode()
    db_execute("UPDATE activos SET foto=? WHERE id=?", (b64,id))
    db_execute("INSERT INTO movimientos (activo_id,tipo,descripcion,usuario) VALUES (?,?,?,?)",
               (id,'Foto','Foto del activo actualizada',session['user']))
    return jsonify({'ok':True,'foto':b64})


@app.route('/api/debug-excel', methods=['POST'])
@admin_required
def debug_excel():
    if 'archivo' not in request.files:
        return jsonify({'error':'No archivo'}), 400
    file = request.files['archivo']
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        hojas = wb.sheetnames
        ws = None
        for name in wb.sheetnames:
            nu = name.upper()
            if 'CARGA' in nu or ('ACTIVO' in nu and 'INSTRUC' not in nu):
                ws = wb[name]; break
        if ws is None:
            ws = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.active
        
        # Leer primeras 6 filas para debug
        preview = []
        for row in range(1, 7):
            fila = []
            for col in range(1, min(ws.max_column+1, 16)):
                v = ws.cell(row=row, column=col).value
                fila.append(str(v) if v is not None else '')
            preview.append(fila)
        
        return jsonify({
            'hojas': hojas,
            'hoja_usada': ws.title,
            'max_row': ws.max_row,
            'max_col': ws.max_column,
            'preview_filas_1_6': preview
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/importar', methods=['POST'])
@admin_required
def importar_excel():
    if 'archivo' not in request.files:
        return jsonify({'error':'No se envio archivo'}), 400
    file = request.files['archivo']
    if not file.filename.lower().endswith(('.xlsx','.xls')):
        return jsonify({'error':'Solo se aceptan archivos Excel (.xlsx)'}), 400
    try:
        import openpyxl
        from datetime import date as dt_date, timedelta as dt_delta

        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)

        # Buscar hoja ACTIVOS
        ws = None
        for name in wb.sheetnames:
            if name.upper() == 'ACTIVOS':
                ws = wb[name]; break
        if ws is None:
            for name in wb.sheetnames:
                if 'LISTA' not in name.upper() and 'REFERENCIA' not in name.upper():
                    ws = wb[name]; break
        if ws is None:
            return jsonify({'error': 'No se encontro hoja ACTIVOS'}), 400

        def leer(row, col):
            v = ws.cell(row=row, column=col).value
            if v is None:
                return ''
            if hasattr(v, 'strftime'):
                return v.strftime('%d-%m-%Y')
            return str(v).strip()

        def leer_fecha(row, col):
            v = ws.cell(row=row, column=col).value
            if v is None:
                return ''
            if isinstance(v, (int, float)) and 30000 < v < 70000:
                try:
                    return (dt_date(1899, 12, 30) + dt_delta(days=int(v))).strftime('%d-%m-%Y')
                except:
                    return ''
            if hasattr(v, 'strftime'):
                return v.strftime('%d-%m-%Y')
            s = str(v).strip().replace('/', '-')
            p = s.split('-')
            if len(p) == 3:
                if len(p[0]) == 4:
                    return f"{p[2].zfill(2)}-{p[1].zfill(2)}-{p[0]}"
                if len(p[2]) == 4:
                    return f"{p[0].zfill(2)}-{p[1].zfill(2)}-{p[2]}"
            return s

        def leer_num(row, col):
            v = ws.cell(row=row, column=col).value
            if v is None: return 0.0
            if isinstance(v, (int, float)): return float(v)
            try: return float(str(v).replace('.','').replace(',','.').replace('$',''))
            except: return 0.0

        def leer_int(row, col, default=7):
            v = ws.cell(row=row, column=col).value
            if v is None: return default
            if isinstance(v, (int, float)): return int(v)
            try: return int(float(str(v)))
            except: return default

        # Leer todos los datos primero
        filas = []
        for row_num in range(4, min(ws.max_row + 1, 600)):
            tipo    = leer(row_num, 1)
            subtipo = leer(row_num, 2)
            if not tipo and not subtipo:
                continue
            filas.append({
                'row': row_num,
                'tipo':     tipo,
                'subtipo':  subtipo,
                'marca':    leer(row_num, 3),
                'modelo':   leer(row_num, 4),
                'serie':    leer(row_num, 5),
                'estado':   leer(row_num, 6) or 'Bueno',
                'edificio': leer(row_num, 7),
                'sala':     leer(row_num, 8),
                'resp':     leer(row_num, 9),
                'cc':       leer(row_num, 10),
                'fecha':    leer_fecha(row_num, 11),
                'precio':   leer_num(row_num, 12),
                'doc':      leer(row_num, 13),
                'vida':     leer_int(row_num, 14, 7),
                'obs':      leer(row_num, 15),
            })

        if not filas:
            return jsonify({'ok': False, 'mensaje': 'No se encontraron datos desde la fila 4', 'errores': []})

        # Abrir UNA sola conexion para todo el proceso
        conn2, mode2 = get_db()
        cur2 = conn2.cursor()

        # Obtener el correlativo actual (maximo en la BD)
        if mode2 == 'pg':
            cur2.execute("SELECT id FROM activos WHERE id LIKE 'AF-%'")
        else:
            cur2.execute("SELECT id FROM activos WHERE id LIKE 'AF-%'")
        existing = cur2.fetchall()

        nums = []
        for row in existing:
            try:
                idd = row[0] if isinstance(row, tuple) else list(row.values())[0]
                parts = str(idd).split('-')
                if len(parts) >= 3:
                    nums.append(int(parts[-1]))
            except:
                pass
        correlativo = max(nums) + 1 if nums else 1000

        creados = 0
        errores = []
        usu = str(session['user'])

        for f in filas:
            if not f['tipo'] or not f['subtipo']:
                errores.append(f"Fila {f['row']}: Tipo o Subtipo vacío")
                continue
            if not f['edificio']:
                errores.append(f"Fila {f['row']}: Edificio vacío")
                continue

            try:
                # Determinar año de la fecha
                fecha = f['fecha']
                anio = datetime.now().year % 100
                if fecha and len(fecha) >= 10:
                    try:
                        anio = datetime.strptime(fecha[:10], '%d-%m-%Y').year % 100
                    except:
                        try:
                            anio = datetime.strptime(fecha[:10], '%Y-%m-%d').year % 100
                        except:
                            pass

                aid = f"AF-{anio:02d}-{correlativo}"
                correlativo += 1

                estado = f['estado']
                if estado not in ('Bueno', 'Regular', 'Malo'):
                    estado = 'Bueno'

                vida = max(1, int(f['vida']))

                if mode2 == 'pg':
                    cur2.execute(
                        "INSERT INTO activos (id,tipo,subtipo,marca,modelo,serie,estado,edificio,sala,responsable,fecha_compra,precio,documento,vida_util,observaciones,foto,centro_costo) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (str(aid), str(f['tipo']), str(f['subtipo']), str(f['marca']),
                         str(f['modelo']), str(f['serie']), str(estado), str(f['edificio']),
                         str(f['sala']), str(f['resp']), str(fecha), float(f['precio']),
                         str(f['doc']), int(vida), str(f['obs']), '', str(f['cc']))
                    )
                    cur2.execute(
                        "INSERT INTO movimientos (activo_id,tipo,descripcion,usuario) VALUES (%s,%s,%s,%s)",
                        (str(aid), 'Alta', f"Importado Excel — {f['subtipo']} {f['marca']} {f['modelo']}", usu)
                    )
                else:
                    cur2.execute(
                        "INSERT INTO activos (id,tipo,subtipo,marca,modelo,serie,estado,edificio,sala,responsable,fecha_compra,precio,documento,vida_util,observaciones,foto,centro_costo) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (str(aid), str(f['tipo']), str(f['subtipo']), str(f['marca']),
                         str(f['modelo']), str(f['serie']), str(estado), str(f['edificio']),
                         str(f['sala']), str(f['resp']), str(fecha), float(f['precio']),
                         str(f['doc']), int(vida), str(f['obs']), '', str(f['cc']))
                    )
                    cur2.execute(
                        "INSERT INTO movimientos (activo_id,tipo,descripcion,usuario) VALUES (?,?,?,?)",
                        (str(aid), 'Alta', f"Importado Excel — {f['subtipo']} {f['marca']} {f['modelo']}", usu)
                    )
                creados += 1

            except Exception as e:
                errores.append(f"Fila {f['row']}: {str(e)}")

        conn2.commit()
        conn2.close()

        msg = f"{creados} activo{'s' if creados!=1 else ''} importado{'s' if creados!=1 else ''} correctamente"
        return jsonify({'ok': True, 'creados': creados, 'errores': errores[:10], 'mensaje': msg})

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'detalle': traceback.format_exc()[-500:]}), 500


@app.route('/api/activos/<id>/qr')
@login_required
def get_qr(id):
    url = request.host_url + f'ficha/{id}'
    try:
        import qrcode
        img = qrcode.make(url)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        b64 = base64.b64encode(buf.getvalue()).decode()
    except:
        b64 = _simple_qr_b64(id)
    return jsonify({'qr':b64,'url':url})

def _simple_qr_b64(text):
    size,cell=21,10; img_size=size*cell+20
    seed=sum(ord(c)*(i+1) for i,c in enumerate(text))
    def dark(r,c):
        if (r<7 and c<7) or (r<7 and c>=size-7) or (r>=size-7 and c<7):
            if r==0 or r==6 or c==0 or c==6: return True
            if 2<=r<=4 and 2<=c<=4: return True
            return False
        return (seed*(r+1)*(c+1)+r*3+c*7)%4!=0
    w=h=img_size; pad=10; rgba=[]
    for y in range(h):
        for x in range(w):
            rc,cc=(y-pad)//cell,(x-pad)//cell
            if 0<=rc<size and 0<=cc<size and dark(rc,cc): rgba+=[0,0,0,255]
            else: rgba+=[255,255,255,255]
    def chunk(name,data):
        c=zlib.crc32(name+data)&0xffffffff
        return struct.pack('>I',len(data))+name+data+struct.pack('>I',c)
    rows=b''
    for y in range(h): rows+=bytes([0])+bytes(rgba[y*w*4:(y+1)*w*4])
    comp=zlib.compress(rows,9)
    ihdr=struct.pack('>II',w,h)+bytes([8,2,0,0,0])
    return base64.b64encode(b'\x89PNG\r\n\x1a\n'+chunk(b'IHDR',ihdr)+chunk(b'IDAT',comp)+chunk(b'IEND',b'')).decode()

@app.route('/ficha/<id>', methods=['GET','POST'])
def ficha_publica(id):
    error = ''
    # Verificar PIN en sesion
    if not session.get('pin_ok'):
        if request.method == 'POST':
            pin = request.form.get('pin','').strip()
            if pin == FICHA_PIN:
                session['pin_ok'] = True
            else:
                error = 'PIN incorrecto, intenta nuevamente'
        if not session.get('pin_ok'):
            return render_template('pin.html', id=id, error=error)
    a = db_fetchone("SELECT * FROM activos WHERE id=?", (id,))
    if not a: return "Activo no encontrado", 404
    dep = calcular_depreciacion(a)
    return render_template('ficha_publica.html', a=a, dep=dep, centros=CENTROS_COSTO, edificios=['Básica','Media','Parvularia','Administración'])

@app.route('/api/stats')
@login_required
def stats():
    total   = db_fetchone("SELECT COUNT(*) as n FROM activos")['n']
    buenos  = db_fetchone("SELECT COUNT(*) as n FROM activos WHERE estado='Bueno'")['n']
    malos   = db_fetchone("SELECT COUNT(*) as n FROM activos WHERE estado='Malo'")['n']
    valor   = db_fetchone("SELECT COALESCE(SUM(precio),0) as s FROM activos")['s']
    por_edificio = db_fetchall("SELECT edificio, COUNT(*) as n FROM activos GROUP BY edificio")
    return jsonify({'total':total,'buenos':buenos,'malos':malos,'valor':valor,'por_edificio':por_edificio})

@app.route('/api/export/excel')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    rows = db_fetchall("SELECT * FROM activos ORDER BY id")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="Activos Fijos"
    headers=['ID Activo','Tipo','Subtipo','Marca','Modelo','N° Serie','Estado',
             'Edificio','Sala','Responsable','Fecha Compra','Precio','Documento','Vida Útil','Observaciones']
    keys=['id','tipo','subtipo','marca','modelo','serie','estado','edificio','sala',
          'responsable','fecha_compra','precio','documento','vida_util','observaciones']
    for col,h in enumerate(headers,1):
        cell=ws.cell(row=1,column=col,value=h)
        cell.font=Font(bold=True,color='FFFFFF')
        cell.fill=PatternFill('solid',start_color='1F3864',fgColor='1F3864')
        cell.alignment=Alignment(horizontal='center')
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=18
    for ri,row in enumerate(rows,2):
        for col,key in enumerate(keys,1):
            ws.cell(row=ri,column=col,value=row.get(key,''))
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,download_name='activos_fijos.xlsx',as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── MANTENCIONES ────────────────────────────────────────────────────────────

@app.route('/api/activos/<id>/mantenciones', methods=['GET'])
@login_required
def get_mantenciones(id):
    rows = db_fetchall(
        "SELECT * FROM mantenciones WHERE activo_id=? ORDER BY fecha DESC", (id,))
    return jsonify(rows)

@app.route('/api/activos/<id>/mantenciones', methods=['POST'])
@admin_required
def crear_mantencion(id):
    d = request.json
    conn2, mode2 = get_db()
    cur2 = conn2.cursor()
    if mode2 == 'pg':
        cur2.execute(
            "INSERT INTO mantenciones (activo_id,fecha,tipo,descripcion,costo,proveedor,estado,proxima_fecha,usuario) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (id, d.get('fecha',''), d.get('tipo','correctiva'),
             d.get('descripcion',''), float(d.get('costo',0) or 0),
             d.get('proveedor',''), d.get('estado','solucionado'),
             d.get('proxima_fecha',''), session['user'])
        )
        cur2.execute(
            "INSERT INTO movimientos (activo_id,tipo,descripcion,usuario) VALUES (%s,%s,%s,%s)",
            (id, 'Mantención',
             f"Mantención {d.get('tipo','correctiva')} — {d.get('descripcion','')} — Costo: ${d.get('costo',0)}",
             session['user'])
        )
    else:
        cur2.execute(
            "INSERT INTO mantenciones (activo_id,fecha,tipo,descripcion,costo,proveedor,estado,proxima_fecha,usuario) VALUES (?,?,?,?,?,?,?,?,?)",
            (id, d.get('fecha',''), d.get('tipo','correctiva'),
             d.get('descripcion',''), float(d.get('costo',0) or 0),
             d.get('proveedor',''), d.get('estado','solucionado'),
             d.get('proxima_fecha',''), session['user'])
        )
        cur2.execute(
            "INSERT INTO movimientos (activo_id,tipo,descripcion,usuario) VALUES (?,?,?,?)",
            (id, 'Mantención',
             f"Mantención {d.get('tipo','correctiva')} — {d.get('descripcion','')} — Costo: ${d.get('costo',0)}",
             session['user'])
        )
    conn2.commit()
    conn2.close()
    return jsonify({'ok': True})

@app.route('/api/mantenciones/<int:mid>', methods=['DELETE'])
@admin_required
def eliminar_mantencion(mid):
    db_execute("DELETE FROM mantenciones WHERE id=?", (mid,))
    return jsonify({'ok': True})

# ─── DASHBOARD / KPIs ────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html', user=session['user'], rol=session['rol'])

@app.route('/api/kpis')
@login_required
def get_kpis():
    tipo     = request.args.get('tipo','')
    subtipo  = request.args.get('subtipo','')
    edificio = request.args.get('edificio','')

    cc_filtro = request.args.get('centro_costo','')
    filtro = " WHERE 1=1"
    params = []
    if tipo:      filtro += " AND tipo=?";          params.append(tipo)
    if subtipo:   filtro += " AND subtipo LIKE ?";  params.append(f'%{subtipo}%')
    if edificio:  filtro += " AND edificio=?";      params.append(edificio)
    if cc_filtro: filtro += " AND centro_costo=?";  params.append(cc_filtro)

    por_estado = db_fetchall(f"SELECT estado, COUNT(*) as n FROM activos{filtro} GROUP BY estado", params)
    por_tipo   = db_fetchall(f"SELECT tipo, COUNT(*) as n FROM activos{filtro} GROUP BY tipo ORDER BY n DESC", params)
    por_edificio=db_fetchall(f"SELECT edificio, COUNT(*) as n FROM activos{filtro} GROUP BY edificio ORDER BY n DESC", params)
    por_cc     =db_fetchall(f"SELECT centro_costo, COUNT(*) as n FROM activos{filtro} AND centro_costo IS NOT NULL AND centro_costo!='' GROUP BY centro_costo ORDER BY n DESC", params)
    totales    = db_fetchone(f"SELECT COUNT(*) as total, COALESCE(SUM(precio),0) as valor FROM activos{filtro}", params)
    # Costo total mantenciones
    costo_mant = db_fetchone(
        "SELECT COALESCE(SUM(costo),0) as total FROM mantenciones")
    # Mantenciones por mes (ultimos 12 meses)
    try:
        mant_mes = db_fetchall(
            """SELECT SUBSTRING(fecha FROM 4 FOR 7) as mes, COUNT(*) as n, COALESCE(SUM(costo),0) as costo
               FROM mantenciones WHERE fecha IS NOT NULL AND fecha != '' GROUP BY mes ORDER BY mes DESC LIMIT 12""")
    except Exception:
        try:
            mant_mes = db_fetchall(
                """SELECT substr(fecha,4,7) as mes, COUNT(*) as n, COALESCE(SUM(costo),0) as costo
                   FROM mantenciones WHERE fecha IS NOT NULL AND fecha != '' GROUP BY mes ORDER BY mes DESC LIMIT 12""")
        except Exception:
            mant_mes = []
    # Activos mas problematicos
    mas_mant = db_fetchall(
        """SELECT a.id, a.subtipo, a.marca, a.modelo, a.edificio,
                  COUNT(m.id) as num_mant, COALESCE(SUM(m.costo),0) as costo_total
           FROM activos a LEFT JOIN mantenciones m ON a.id=m.activo_id
           GROUP BY a.id, a.subtipo, a.marca, a.modelo, a.edificio
           HAVING COUNT(m.id) > 0
           ORDER BY COUNT(m.id) DESC, COALESCE(SUM(m.costo),0) DESC LIMIT 10""")
    # Costo por tipo de activo
    costo_tipo = db_fetchall(
        """SELECT a.tipo, COALESCE(SUM(m.costo),0) as costo_total, COUNT(m.id) as num_mant
           FROM activos a LEFT JOIN mantenciones m ON a.id=m.activo_id
           GROUP BY a.tipo ORDER BY costo_total DESC""")
    # Depreciacion total
    activos_dep = db_fetchall(
        "SELECT tipo, precio, fecha_compra, vida_util FROM activos WHERE precio > 0 AND fecha_compra != ''")
    valor_actual_total = 0
    dep_acum_total = 0
    proximos_depreciar = []
    from datetime import datetime
    for a in activos_dep:
        try:
            fecha = None
            for fmt in ['%d-%m-%Y','%Y-%m-%d','%d/%m/%Y']:
                try: fecha = datetime.strptime(str(a['fecha_compra'])[:10], fmt); break
                except: pass
            if not fecha: continue
            precio = float(a['precio'])
            vida = int(a['vida_util'] or 7)
            tasa = 1.0 / vida
            anos = (datetime.now() - fecha).days / 365.25
            val_res = precio * 0.10
            dep = min(precio - val_res, (precio - val_res) * tasa * anos)
            val_act = max(val_res, precio - dep)
            valor_actual_total += val_act
            dep_acum_total += dep
            pct = min(100, dep / (precio - val_res) * 100) if precio > val_res else 100
            if pct >= 70:
                proximos_depreciar.append({
                    'tipo': a['tipo'], 'porcentaje': round(pct,1),
                    'valor_actual': round(val_act)
                })
        except: pass

    return jsonify({
        'por_estado':        por_estado,
        'por_cc':            por_cc,
        'por_tipo':          por_tipo,
        'por_edificio':      por_edificio,
        'totales':           totales,
        'costo_mantenciones':costo_mant['total'] if costo_mant else 0,
        'mant_por_mes':      list(reversed(mant_mes)),
        'mas_problematicos': mas_mant,
        'costo_por_tipo':    costo_tipo,
        'valor_actual_total':round(valor_actual_total),
        'dep_acum_total':    round(dep_acum_total),
        'proximos_depreciar':sorted(proximos_depreciar, key=lambda x:-x['porcentaje'])[:5],
    })


@app.route('/api/kpis/subtipo')
@login_required
def kpis_subtipo():
    tipo     = request.args.get('tipo','')
    subtipo  = request.args.get('subtipo','')
    edificio = request.args.get('edificio','')
    filtro = " WHERE subtipo IS NOT NULL AND subtipo != ''"
    params = []
    if tipo:     filtro += " AND tipo=?";         params.append(tipo)
    if subtipo:  filtro += " AND subtipo LIKE ?";  params.append(f'%{subtipo}%')
    if edificio: filtro += " AND edificio=?";      params.append(edificio)

    por_subtipo = db_fetchall(
        f"SELECT subtipo, COUNT(*) as n FROM activos{filtro} GROUP BY subtipo ORDER BY n DESC", params)
    filtro_ed = filtro + " AND edificio IS NOT NULL AND edificio != ''"
    por_subtipo_edificio = db_fetchall(
        f"SELECT subtipo, edificio, COUNT(*) as n FROM activos{filtro_ed} GROUP BY subtipo, edificio ORDER BY subtipo, n DESC", params)
    return jsonify({
        'por_subtipo': por_subtipo,
        'por_subtipo_edificio': por_subtipo_edificio
    })


# ── MOVIMIENTOS DE ACTIVOS ────────────────────────────────────────────────────

@app.route('/movimientos')
@login_required
def movimientos_page():
    return render_template('movimientos.html', user=session['user'], rol=session['rol'])

@app.route('/api/movimientos/activos', methods=['GET'])
@login_required
def get_movimientos_activos():
    tipo_mov  = request.args.get('tipo_mov','')
    q         = request.args.get('q','')
    tipo_activo = request.args.get('tipo_activo','')
    edificio  = request.args.get('edificio','')
    desde     = request.args.get('desde','')
    hasta     = request.args.get('hasta','')

    sql = """SELECT m.id, m.activo_id, m.tipo as tipo_mov, m.descripcion,
                    m.usuario, m.fecha, m.responsable, m.edificio,
                    a.tipo as tipo_activo, a.subtipo, a.marca, a.modelo
             FROM movimientos_activos m
             LEFT JOIN activos a ON m.activo_id=a.id
             WHERE 1=1"""
    params = []
    if tipo_mov:     sql += " AND m.tipo=?";           params.append(tipo_mov)
    if q:            sql += " AND (m.activo_id LIKE ? OR a.subtipo LIKE ? OR a.marca LIKE ?)"; params+=[f'%{q}%']*3
    if tipo_activo:  sql += " AND a.tipo=?";           params.append(tipo_activo)
    if edificio:     sql += " AND m.edificio=?";       params.append(edificio)
    if desde:
        p=desde.split('-'); desde_fmt=f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else desde
        sql += " AND m.fecha >= ?"; params.append(desde_fmt)
    if hasta:
        p=hasta.split('-'); hasta_fmt=f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else hasta
        sql += " AND m.fecha <= ?"; params.append(hasta_fmt)
    sql += " ORDER BY m.created_at DESC LIMIT 500"
    return jsonify(db_fetchall(sql, params))

@app.route('/api/movimientos/activos', methods=['POST'])
@login_required
def crear_movimiento_activo():
    d = request.json
    aid = d.get('activo_id','').strip().upper()
    if not aid: return jsonify({'error':'ID activo requerido'}), 400
    a = db_fetchone("SELECT * FROM activos WHERE id=?", (aid,))
    if not a: return jsonify({'error':f'Activo {aid} no encontrado'}), 404

    fecha = d.get('fecha', datetime.now().strftime('%d-%m-%Y'))
    if fecha and '-' in fecha:
        parts = fecha.split('-')
        if len(parts[0])==4:
            fecha = f"{parts[2]}-{parts[1]}-{parts[0]}"

    conn2, mode2 = get_db()
    cur2 = conn2.cursor()
    if mode2 == 'pg':
        cur2.execute(
            """INSERT INTO movimientos_activos
               (activo_id,tipo,descripcion,edificio,responsable,usuario,fecha)
               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (aid, d.get('tipo_mov','entrada'), d.get('descripcion',''),
             d.get('edificio',''), d.get('responsable',''),
             session['user'], fecha))
    else:
        cur2.execute(
            """INSERT INTO movimientos_activos
               (activo_id,tipo,descripcion,edificio,responsable,usuario,fecha)
               VALUES (?,?,?,?,?,?,?)""",
            (aid, d.get('tipo_mov','entrada'), d.get('descripcion',''),
             d.get('edificio',''), d.get('responsable',''),
             session['user'], fecha))
    # Si es salida, cambiar estado del activo a "De Baja"
    if d.get('tipo_mov') == 'salida':
        if mode2 == 'pg':
            cur2.execute("UPDATE activos SET estado='De Baja' WHERE id=%s", (aid,))
        else:
            cur2.execute("UPDATE activos SET estado='De Baja' WHERE id=?", (aid,))

    conn2.commit(); conn2.close()
    return jsonify({'ok':True})

@app.route('/api/movimientos/activos/<int:mid>', methods=['DELETE'])
@login_required
def eliminar_movimiento_activo(mid):
    # Obtener el movimiento antes de borrar
    mov = db_fetchone("SELECT * FROM movimientos_activos WHERE id=?", (mid,))
    db_execute("DELETE FROM movimientos_activos WHERE id=?", (mid,))
    # Si era salida y no hay otras guías de salida para ese activo, revertir estado
    if mov and mov.get('tipo') == 'salida':
        otras = db_fetchall(
            "SELECT id FROM movimientos_activos WHERE activo_id=? AND tipo='salida'",
            (mov['activo_id'],))
        if not otras:
            db_execute("UPDATE activos SET estado='Bueno' WHERE id=? AND estado='De Baja'",
                      (mov['activo_id'],))
    return jsonify({'ok':True})

@app.route('/api/export/movimientos-activos')
@login_required
def export_movimientos_activos():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    rows = db_fetchall(
        """SELECT m.fecha, m.tipo, m.activo_id, m.descripcion,
                  a.tipo as tipo_activo, a.subtipo, a.marca, a.modelo,
                  m.edificio, m.responsable, m.usuario
           FROM movimientos_activos m LEFT JOIN activos a ON m.activo_id=a.id
           ORDER BY m.created_at DESC""")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="Movimientos Activos"
    headers=['Fecha','Tipo','ID Activo','Descripción','Tipo Activo','Subtipo','Marca','Modelo','Edificio','Responsable','Usuario']
    for col,h in enumerate(headers,1):
        cell=ws.cell(row=1,column=col,value=h)
        cell.font=Font(bold=True,color='FFFFFF')
        cell.fill=PatternFill('solid',start_color='1F3864',fgColor='1F3864')
        cell.alignment=Alignment(horizontal='center')
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=18
    for ri,r in enumerate(rows,2):
        for col,key in enumerate(['fecha','tipo','activo_id','descripcion','tipo_activo','subtipo','marca','modelo','edificio','responsable','usuario'],1):
            ws.cell(row=ri,column=col,value=r.get(key,''))
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,download_name='movimientos_activos.xlsx',as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/export/informe-anual')
@login_required
def export_informe_anual():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    anio = request.args.get('anio', str(datetime.now().year))
    wb = openpyxl.Workbook()

    AZUL_OSC = "1F3864"
    AZUL_MED = "2E74B5"
    AZUL_SUV = "E6F1FB"
    VERDE    = "E2EFDA"
    VERDE_OSC= "27500A"
    AMARILLO = "FFF2CC"
    ROJO_SUV = "FCEBEB"
    BLANCO   = "FFFFFF"
    GRIS     = "F2F2F2"

    def fill(c):  return PatternFill("solid", start_color=c, fgColor=c)
    def borde():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)
    def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def hdr(ws, col, row, val, bg=None, bold=True, size=11, color="FFFFFF", aln=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name="Arial", bold=bold, size=size, color=color)
        if bg: cell.fill = fill(bg)
        cell.alignment = aln or center()
        cell.border = borde()
        return cell

    def dat(ws, col, row, val, bg=None, bold=False, color="000000", aln=None, fmt=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name="Arial", bold=bold, size=10, color=color)
        if bg: cell.fill = fill(bg)
        cell.alignment = aln or center()
        cell.border = borde()
        if fmt: cell.number_format = fmt
        return cell

    # ── HOJA 1: PORTADA ──────────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "Portada"
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 5
    ws0.column_dimensions["B"].width = 60

    ws0.merge_cells("B2:B3")
    c = ws0["B2"]
    c.value = f"INFORME ANUAL DE ACTIVOS FIJOS"
    c.font = Font(name="Arial", bold=True, size=20, color=BLANCO)
    c.fill = fill(AZUL_OSC)
    c.alignment = center()
    ws0.row_dimensions[2].height = 40
    ws0.row_dimensions[3].height = 40

    ws0.merge_cells("B4:B4")
    c = ws0["B4"]
    c.value = f"Año {anio} — Colegio"
    c.font = Font(name="Arial", size=14, color=BLANCO, italic=True)
    c.fill = fill(AZUL_MED)
    c.alignment = center()
    ws0.row_dimensions[4].height = 28

    ws0.merge_cells("B5:B5")
    c = ws0["B5"]
    c.value = f"Generado el {datetime.now().strftime('%d-%m-%Y a las %H:%M')}"
    c.font = Font(name="Arial", size=11, color="555555")
    c.fill = fill(GRIS)
    c.alignment = center()
    ws0.row_dimensions[5].height = 22

    # Resumen ejecutivo
    activos = db_fetchall("SELECT * FROM activos")
    total   = len(activos)
    buenos  = sum(1 for a in activos if a.get('estado')=='Bueno')
    regulares=sum(1 for a in activos if a.get('estado')=='Regular')
    malos   = sum(1 for a in activos if a.get('estado')=='Malo')
    valor_total = sum(float(a.get('precio') or 0) for a in activos)

    # Depreciacion
    valor_actual_total = 0
    dep_total = 0
    for a in activos:
        try:
            fecha = None
            for fmt in ['%d-%m-%Y','%Y-%m-%d']:
                try: fecha = datetime.strptime(str(a['fecha_compra'])[:10], fmt); break
                except: pass
            if not fecha: continue
            precio = float(a['precio'] or 0)
            vida   = int(a['vida_util'] or 7)
            tasa   = 1.0/vida
            anos   = (datetime.now()-fecha).days/365.25
            val_res= precio*0.10
            dep    = min(precio-val_res,(precio-val_res)*tasa*anos)
            val_act= max(val_res, precio-dep)
            valor_actual_total += val_act
            dep_total += dep
        except: pass

    mant_total = db_fetchone("SELECT COALESCE(SUM(costo),0) as s FROM mantenciones")
    bajas = db_fetchall("SELECT COUNT(*) as n FROM movimientos_activos WHERE tipo='salida'")

    resumen = [
        ("📦 Total activos registrados",   total,          AZUL_SUV),
        ("✅ Activos en buen estado",       buenos,         VERDE),
        ("⚠️ Activos en estado regular",   regulares,      AMARILLO),
        ("❌ Activos en mal estado",        malos,          ROJO_SUV),
        ("💰 Valor bruto inventario",       f"${round(valor_total):,}".replace(',','.'), AZUL_SUV),
        ("📉 Valor actual (depreciado)",    f"${round(valor_actual_total):,}".replace(',','.'), AZUL_SUV),
        ("⬇️ Depreciación acumulada",      f"${round(dep_total):,}".replace(',','.'),   AMARILLO),
        ("🔧 Gasto total mantenciones",     f"${round(mant_total['s'] if mant_total else 0):,}".replace(',','.'), ROJO_SUV),
        ("📤 Guías de salida emitidas",     bajas[0]['n'] if bajas else 0, GRIS),
    ]

    row = 7
    ws0.merge_cells(f"B{row}:B{row}")
    c = ws0[f"B{row}"]
    c.value = "RESUMEN EJECUTIVO"
    c.font = Font(name="Arial", bold=True, size=12, color=BLANCO)
    c.fill = fill(AZUL_OSC)
    c.alignment = center()
    ws0.row_dimensions[row].height = 24
    row += 1

    for label, valor, bg in resumen:
        ws0.row_dimensions[row].height = 20
        c1 = ws0[f"B{row}"]
        c1.value = f"  {label}"
        c1.font  = Font(name="Arial", size=11, bold=False)
        c1.fill  = fill(bg)
        c1.alignment = left()
        c1.border = borde()
        # Valor en misma celda como segunda columna
        ws0.merge_cells(f"B{row}:B{row}")
        c1.value = f"{label}:   {valor}"
        row += 1

    # ── HOJA 2: INVENTARIO COMPLETO ──────────────────────────────────────────
    ws1 = wb.create_sheet("Inventario Completo")
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A3"

    cols_inv = [
        ("ID Activo",20),("Tipo",22),("Subtipo",18),("Marca",14),("Modelo",18),
        ("N° Serie",18),("Estado",12),("Edificio",14),("Sala",14),("Responsable",20),
        ("Fecha Compra",14),("Precio",14),("Vida Útil",10),("Valor Actual",14),("Dep. Acum.",14)
    ]
    ws1.merge_cells(f"A1:{get_column_letter(len(cols_inv))}1")
    c=ws1["A1"]
    c.value=f"INVENTARIO COMPLETO DE ACTIVOS FIJOS — Año {anio}"
    c.font=Font(name="Arial",bold=True,size=13,color=BLANCO)
    c.fill=fill(AZUL_OSC)
    c.alignment=center()
    ws1.row_dimensions[1].height=28

    for col,(hd,w) in enumerate(cols_inv,1):
        hdr(ws1,col,2,hd,bg=AZUL_MED)
        ws1.column_dimensions[get_column_letter(col)].width=w
    ws1.row_dimensions[2].height=22

    for ri,a in enumerate(activos,3):
        bg = BLANCO if ri%2==0 else GRIS
        # Calcular valor actual y dep
        val_act=0; dep_a=0
        try:
            fecha=None
            for fmt in ['%d-%m-%Y','%Y-%m-%d']:
                try: fecha=datetime.strptime(str(a['fecha_compra'])[:10],fmt); break
                except: pass
            if fecha and a.get('precio'):
                precio=float(a['precio']); vida=int(a['vida_util'] or 7)
                anos=(datetime.now()-fecha).days/365.25
                val_res=precio*0.10; dep=min(precio-val_res,(precio-val_res)/vida*anos)
                val_act=max(val_res,precio-dep); dep_a=dep
        except: pass

        estado_bg = VERDE if a.get('estado')=='Bueno' else AMARILLO if a.get('estado')=='Regular' else ROJO_SUV
        vals = [a.get('id'),a.get('tipo'),a.get('subtipo'),a.get('marca'),a.get('modelo'),
                a.get('serie'),a.get('estado'),a.get('edificio'),a.get('sala'),a.get('responsable'),
                a.get('fecha_compra'),float(a.get('precio') or 0),a.get('vida_util'),round(val_act),round(dep_a)]
        for col,val in enumerate(vals,1):
            b = estado_bg if col==7 else bg
            cell=dat(ws1,col,ri,val,bg=b)
            if col in [12,14,15]: cell.number_format='#,##0'
        ws1.row_dimensions[ri].height=16

    # ── HOJA 3: RESUMEN POR EDIFICIO ─────────────────────────────────────────
    ws2 = wb.create_sheet("Por Edificio")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:F1")
    c=ws2["A1"]
    c.value=f"RESUMEN POR EDIFICIO — Año {anio}"
    c.font=Font(name="Arial",bold=True,size=13,color=BLANCO)
    c.fill=fill(AZUL_OSC); c.alignment=center()
    ws2.row_dimensions[1].height=28

    edificios_data = db_fetchall(
        """SELECT edificio, COUNT(*) as n,
                  SUM(CASE WHEN estado='Bueno' THEN 1 ELSE 0 END) as buenos,
                  SUM(CASE WHEN estado='Regular' THEN 1 ELSE 0 END) as regulares,
                  SUM(CASE WHEN estado='Malo' THEN 1 ELSE 0 END) as malos,
                  COALESCE(SUM(precio),0) as valor
           FROM activos GROUP BY edificio ORDER BY n DESC""")

    hdrs_ed = ["Edificio","Total Activos","Bueno","Regular","Malo","Valor Inventario"]
    for col,h in enumerate(hdrs_ed,1):
        hdr(ws2,col,2,h,bg=AZUL_MED)
        ws2.column_dimensions[get_column_letter(col)].width=20
    ws2.row_dimensions[2].height=22

    for ri,ed in enumerate(edificios_data,3):
        bg = BLANCO if ri%2==0 else GRIS
        ws2.row_dimensions[ri].height=18
        dat(ws2,1,ri,ed.get('edificio','Sin asignar'),bg=bg,bold=True,aln=left())
        dat(ws2,2,ri,ed.get('n',0),bg=bg)
        dat(ws2,3,ri,ed.get('buenos',0),bg=VERDE)
        dat(ws2,4,ri,ed.get('regulares',0),bg=AMARILLO)
        dat(ws2,5,ri,ed.get('malos',0),bg=ROJO_SUV)
        dat(ws2,6,ri,float(ed.get('valor',0)),bg=bg,fmt='#,##0')

    # ── HOJA 4: RESUMEN POR TIPO Y SUBTIPO ───────────────────────────────────
    ws3 = wb.create_sheet("Por Tipo y Subtipo")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:E1")
    c=ws3["A1"]
    c.value=f"DISTRIBUCIÓN POR TIPO Y SUBTIPO — Año {anio}"
    c.font=Font(name="Arial",bold=True,size=13,color=BLANCO)
    c.fill=fill(AZUL_OSC); c.alignment=center()
    ws3.row_dimensions[1].height=28

    tipo_data = db_fetchall(
        """SELECT tipo, subtipo, COUNT(*) as n,
                  SUM(CASE WHEN estado='Bueno' THEN 1 ELSE 0 END) as buenos,
                  COALESCE(SUM(precio),0) as valor
           FROM activos GROUP BY tipo, subtipo ORDER BY tipo, n DESC""")

    hdrs_t = ["Tipo","Subtipo","Cantidad","En buen estado","Valor"]
    for col,h in enumerate(hdrs_t,1):
        hdr(ws3,col,2,h,bg=AZUL_MED)
        ws3.column_dimensions[get_column_letter(col)].width=[22,18,12,16,16][col-1]
    ws3.row_dimensions[2].height=22

    prev_tipo=""
    for ri,t in enumerate(tipo_data,3):
        bg = AZUL_SUV if t.get('tipo')!=prev_tipo else (BLANCO if ri%2==0 else GRIS)
        bold_tipo = t.get('tipo')!=prev_tipo
        ws3.row_dimensions[ri].height=16
        dat(ws3,1,ri,t.get('tipo',''),bg=bg,bold=bold_tipo,aln=left())
        dat(ws3,2,ri,t.get('subtipo',''),bg=bg,aln=left())
        dat(ws3,3,ri,t.get('n',0),bg=bg)
        dat(ws3,4,ri,t.get('buenos',0),bg=VERDE if t.get('buenos')==t.get('n') else bg)
        dat(ws3,5,ri,float(t.get('valor',0)),bg=bg,fmt='#,##0')
        prev_tipo=t.get('tipo','')

    # ── HOJA 5: MANTENCIONES ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Mantenciones")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:H1")
    c=ws4["A1"]
    c.value=f"REGISTRO DE MANTENCIONES — Año {anio}"
    c.font=Font(name="Arial",bold=True,size=13,color=BLANCO)
    c.fill=fill(AZUL_OSC); c.alignment=center()
    ws4.row_dimensions[1].height=28

    mant_data = db_fetchall(
        """SELECT m.fecha, m.tipo, m.descripcion, m.costo, m.proveedor, m.estado,
                  m.activo_id, a.subtipo, a.edificio
           FROM mantenciones m LEFT JOIN activos a ON m.activo_id=a.id
           ORDER BY m.fecha DESC""")

    hdrs_m=["Fecha","ID Activo","Subtipo","Edificio","Tipo","Descripción","Costo","Proveedor","Estado"]
    for col,h in enumerate(hdrs_m,1):
        hdr(ws4,col,2,h,bg=AZUL_MED)
        ws4.column_dimensions[get_column_letter(col)].width=[12,14,16,14,12,30,12,20,12][col-1]
    ws4.row_dimensions[2].height=22

    for ri,m in enumerate(mant_data,3):
        bg=BLANCO if ri%2==0 else GRIS
        ws4.row_dimensions[ri].height=16
        dat(ws4,1,ri,m.get('fecha',''),bg=bg)
        dat(ws4,2,ri,m.get('activo_id',''),bg=bg,bold=True,color=AZUL_OSC)
        dat(ws4,3,ri,m.get('subtipo',''),bg=bg)
        dat(ws4,4,ri,m.get('edificio',''),bg=bg)
        dat(ws4,5,ri,m.get('tipo',''),bg=bg)
        dat(ws4,6,ri,m.get('descripcion',''),bg=bg,aln=left())
        dat(ws4,7,ri,float(m.get('costo') or 0),bg=bg,fmt='#,##0')
        dat(ws4,8,ri,m.get('proveedor',''),bg=bg,aln=left())
        estado_bg=VERDE if m.get('estado')=='solucionado' else AMARILLO
        dat(ws4,9,ri,m.get('estado',''),bg=estado_bg)

    # Totales mantenciones
    if mant_data:
        tot_row = len(mant_data)+3
        ws4.row_dimensions[tot_row].height=20
        c=ws4.cell(row=tot_row,column=6,value="TOTAL GASTO MANTENCIONES:")
        c.font=Font(name="Arial",bold=True,size=11)
        c.fill=fill(AZUL_SUV); c.alignment=center(); c.border=borde()
        tot_mant=sum(float(m.get('costo') or 0) for m in mant_data)
        c2=ws4.cell(row=tot_row,column=7,value=round(tot_mant))
        c2.font=Font(name="Arial",bold=True,size=11,color="C00000")
        c2.fill=fill(ROJO_SUV); c2.alignment=center()
        c2.number_format='#,##0'; c2.border=borde()

    # ── HOJA 6: GUÍAS DE SALIDA ───────────────────────────────────────────────
    ws5 = wb.create_sheet("Guías de Salida")
    ws5.sheet_view.showGridLines = False

    ws5.merge_cells("A1:G1")
    c=ws5["A1"]
    c.value=f"GUÍAS DE SALIDA — BAJAS Y SALIDAS DE INVENTARIO — Año {anio}"
    c.font=Font(name="Arial",bold=True,size=13,color=BLANCO)
    c.fill=fill(AZUL_OSC); c.alignment=center()
    ws5.row_dimensions[1].height=28

    try:
        bajas_data = db_fetchall(
            """SELECT m.fecha, m.activo_id, m.descripcion, m.edificio,
                      m.responsable, m.usuario, a.subtipo, a.marca, a.modelo, a.precio
               FROM movimientos_activos m LEFT JOIN activos a ON m.activo_id=a.id
               WHERE m.tipo='salida' ORDER BY m.fecha DESC""")
    except: bajas_data=[]

    hdrs_b=["Fecha","ID Activo","Subtipo","Marca/Modelo","Descripción/Motivo","Edificio","Responsable","Valor Original","Usuario"]
    for col,h in enumerate(hdrs_b,1):
        hdr(ws5,col,2,h,bg=AZUL_MED)
        ws5.column_dimensions[get_column_letter(col)].width=[12,14,16,18,28,14,18,14,14][col-1]
    ws5.row_dimensions[2].height=22

    for ri,b in enumerate(bajas_data,3):
        bg=BLANCO if ri%2==0 else GRIS
        ws5.row_dimensions[ri].height=16
        dat(ws5,1,ri,b.get('fecha',''),bg=bg)
        dat(ws5,2,ri,b.get('activo_id',''),bg=bg,bold=True,color="C00000")
        dat(ws5,3,ri,b.get('subtipo',''),bg=bg)
        dat(ws5,4,ri,f"{b.get('marca','')} {b.get('modelo','')}".strip(),bg=bg)
        dat(ws5,5,ri,b.get('descripcion',''),bg=bg,aln=left())
        dat(ws5,6,ri,b.get('edificio',''),bg=bg)
        dat(ws5,7,ri,b.get('responsable',''),bg=bg)
        dat(ws5,8,ri,float(b.get('precio') or 0),bg=ROJO_SUV,fmt='#,##0')
        dat(ws5,9,ri,b.get('usuario',''),bg=bg)

    if not bajas_data:
        ws5.merge_cells("A3:I3")
        c=ws5["A3"]
        c.value="Sin guías de salida registradas"
        c.font=Font(name="Arial",italic=True,color="999999")
        c.alignment=center()

    # ── HOJA 7: DEPRECIACIÓN ─────────────────────────────────────────────────
    ws6 = wb.create_sheet("Depreciación")
    ws6.sheet_view.showGridLines = False

    ws6.merge_cells("A1:H1")
    c=ws6["A1"]
    c.value=f"TABLA DE DEPRECIACIÓN SII — Año {anio}"
    c.font=Font(name="Arial",bold=True,size=13,color=BLANCO)
    c.fill=fill(AZUL_OSC); c.alignment=center()
    ws6.row_dimensions[1].height=28

    hdrs_d=["ID Activo","Tipo","Subtipo","Fecha Compra","Precio Original","Vida Útil","Años Uso","Valor Actual","Dep. Acumulada","% Depreciado"]
    for col,h in enumerate(hdrs_d,1):
        hdr(ws6,col,2,h,bg=AZUL_MED)
        ws6.column_dimensions[get_column_letter(col)].width=[14,20,16,14,16,10,10,14,14,12][col-1]
    ws6.row_dimensions[2].height=22

    activos_dep = [a for a in activos if a.get('precio') and a.get('fecha_compra')]
    for ri,a in enumerate(activos_dep,3):
        try:
            fecha=None
            for fmt in ['%d-%m-%Y','%Y-%m-%d']:
                try: fecha=datetime.strptime(str(a['fecha_compra'])[:10],fmt); break
                except: pass
            if not fecha: continue
            precio=float(a['precio']); vida=int(a['vida_util'] or 7)
            anos=round((datetime.now()-fecha).days/365.25,1)
            val_res=precio*0.10
            dep=min(precio-val_res,(precio-val_res)/vida*anos)
            val_act=max(val_res,precio-dep)
            pct=min(100,round(dep/(precio-val_res)*100,1)) if precio>val_res else 100
            bg=ROJO_SUV if pct>=80 else AMARILLO if pct>=50 else BLANCO if ri%2==0 else GRIS
            ws6.row_dimensions[ri].height=16
            dat(ws6,1,ri,a.get('id'),bg=bg,bold=True,color=AZUL_OSC)
            dat(ws6,2,ri,a.get('tipo'),bg=bg,aln=left())
            dat(ws6,3,ri,a.get('subtipo'),bg=bg,aln=left())
            dat(ws6,4,ri,a.get('fecha_compra'),bg=bg)
            dat(ws6,5,ri,precio,bg=bg,fmt='#,##0')
            dat(ws6,6,ri,f"{vida} años",bg=bg)
            dat(ws6,7,ri,f"{anos} años",bg=bg)
            dat(ws6,8,ri,round(val_act),bg=bg,fmt='#,##0')
            dat(ws6,9,ri,round(dep),bg=bg,fmt='#,##0')
            c=dat(ws6,10,ri,f"{pct}%",bg=bg,bold=pct>=80)
            if pct>=80: c.font=Font(name="Arial",bold=True,color="C00000",size=10)
        except: pass

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"Informe_Activos_Fijos_{anio}.xlsx"
    return send_file(buf, download_name=fname, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/activos/<id>/historial')
def get_historial_publico(id):
    if not session.get('pin_ok') and 'user' not in session:
        return jsonify([])
    movs = db_fetchall(
        "SELECT tipo, descripcion, usuario, fecha FROM movimientos WHERE activo_id=? ORDER BY id DESC LIMIT 30",
        (id,))
    return jsonify(movs)

if __name__=='__main__':
    init_db()
    app.run(debug=False,host='0.0.0.0',port=int(os.environ.get('PORT',5000)))
